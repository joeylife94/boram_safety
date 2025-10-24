from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request
from sqlalchemy.orm import Session
from typing import List, Optional
import math

from database import get_db
from crud import product as product_crud
from crud import category as category_crud
from crud import audit as audit_crud
from schemas.product import ProductResponse, ProductCreate, ProductUpdate, ProductSearchParams, ProductSearchResponse
from schemas.category import Category, CategoryCreate, CategoryUpdate
from schemas.audit import AuditLogResponse, AuditLogFilter
from models.safety import SafetyProduct, SafetyCategory
from models.audit import AuditAction, AuditEntityType
from core.logger import get_logger
from utils.audit_logger import (
    log_product_create, log_product_update, log_product_delete,
    log_category_create, log_category_update, log_category_delete,
    log_bulk_update, log_bulk_delete, model_to_dict
)

logger = get_logger(__name__)

# 🔐 Admin Router - 모든 CRUD 작업 허용
router = APIRouter(
    tags=["admin-api"]
)

@router.get("/health")
def admin_health_check():
    """Admin API 상태 확인"""
    return {"status": "healthy", "role": "admin"}

@router.get("/dashboard")
async def admin_dashboard(db: Session = Depends(get_db)):
    """관리자 대시보드 통계 정보를 반환합니다."""
    from sqlalchemy import func, and_
    from datetime import datetime, timedelta
    
    # 기본 통계
    total_products = product_crud.get_product_count(db)
    total_categories = category_crud.get_category_count(db)
    featured_products = product_crud.get_featured_product_count(db)
    
    # 카테고리별 제품 수
    category_stats = db.query(
        SafetyCategory.name,
        SafetyCategory.code,
        func.count(SafetyProduct.id).label('product_count')
    ).outerjoin(SafetyProduct, SafetyCategory.id == SafetyProduct.category_id)\
     .group_by(SafetyCategory.id, SafetyCategory.name, SafetyCategory.code)\
     .order_by(func.count(SafetyProduct.id).desc())\
     .all()
    
    # 최근 7일간 등록된 제품
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_products = db.query(SafetyProduct)\
        .filter(SafetyProduct.created_at >= seven_days_ago)\
        .order_by(SafetyProduct.created_at.desc())\
        .limit(10)\
        .all()
    
    # 최근 수정된 제품
    recently_updated = db.query(SafetyProduct)\
        .filter(SafetyProduct.updated_at.isnot(None))\
        .order_by(SafetyProduct.updated_at.desc())\
        .limit(10)\
        .all()
    
    # 이미지 없는 제품 (file_path가 default인 경우)
    products_without_images = db.query(SafetyProduct)\
        .filter(SafetyProduct.file_path.like('%default%'))\
        .count()
    
    # 재고 부족 제품 (stock_status가 'out_of_stock'인 경우)
    out_of_stock_count = db.query(SafetyProduct)\
        .filter(SafetyProduct.stock_status == 'out_of_stock')\
        .count()
    
    return {
        "summary": {
            "total_products": total_products,
            "total_categories": total_categories,
            "featured_products": featured_products,
            "out_of_stock": out_of_stock_count,
            "missing_images": products_without_images
        },
        "category_stats": [
            {
                "name": stat.name,
                "code": stat.code,
                "count": stat.product_count
            } for stat in category_stats
        ],
        "recent_products": [
            {
                "id": p.id,
                "name": p.name,
                "model_number": p.model_number,
                "created_at": p.created_at.isoformat() if p.created_at else None
            } for p in recent_products
        ],
        "recently_updated": [
            {
                "id": p.id,
                "name": p.name,
                "model_number": p.model_number,
                "updated_at": p.updated_at.isoformat() if p.updated_at else None
            } for p in recently_updated
        ]
    }

@router.post("/upload-image")
async def upload_image(file: UploadFile = File(...)):
    """제품 이미지를 업로드합니다."""
    import os
    import uuid
    from pathlib import Path
    
    try:
        # 허용된 파일 형식 체크
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif'}
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(status_code=400, detail="지원하지 않는 파일 형식입니다")
        
        # 고유한 파일명 생성
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        
        # 업로드 디렉토리 설정 (public/images/)
        upload_dir = Path("../frontend/public/images")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # 파일 저장
        file_path = upload_dir / unique_filename
        
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
        
        # 웹에서 접근 가능한 URL 반환
        file_url = f"/images/{unique_filename}"
        
        return {"url": file_url, "filename": unique_filename}
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 업로드 중 오류가 발생했습니다: {str(e)}")

# ============= 카테고리 관리 =============

@router.get("/categories", response_model=List[Category])
async def read_categories(
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db)
):
    """카테고리 목록을 조회합니다."""
    return category_crud.get_categories(db, skip=skip, limit=limit)

@router.get("/categories/{category_id}", response_model=Category)
async def read_category(
    category_id: int,
    db: Session = Depends(get_db)
):
    """특정 카테고리를 조회합니다."""
    category = category_crud.get_category(db, category_id)
    if category is None:
        raise HTTPException(status_code=404, detail="Category not found")
    return category

@router.post("/categories", response_model=Category)
async def create_category(
    category: CategoryCreate,
    request: Request,
    db: Session = Depends(get_db)
):
    """새로운 카테고리를 생성합니다 (JSON 지원)."""
    logger.info(f"Creating category: {category.dict()}")
    created_category = category_crud.create_category(db, category)
    
    # Audit Log 기록
    log_category_create(db, created_category.id, category.dict(), request)
    
    return created_category

@router.post("/categories/form", response_model=Category)
async def create_category_form(
    name: str = Form(...),
    code: str = Form(...),
    slug: str = Form(...),
    description: Optional[str] = Form(None),
    display_order: Optional[int] = Form(0),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    """새로운 카테고리를 생성합니다 (FormData 지원)."""
    from schemas.category import CategoryCreate
    
    # CategoryCreate 객체 생성
    category_data = {
        'name': name.strip(),
        'code': code.strip(),
        'slug': slug.strip(),
        'description': description.strip() if description and description.strip() else None,
        'display_order': display_order or 0
    }
    
    category = CategoryCreate(**category_data)
    created_category = category_crud.create_category(db, category)
    
    return created_category

@router.put("/categories/{category_id}", response_model=Category)
async def update_category(
    category_id: int,
    category: CategoryUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """카테고리 정보를 수정합니다 (JSON 지원)."""
    logger.info(f"Updating category {category_id}: {category.dict(exclude_unset=True)}")
    
    # 기존 데이터 조회 (Audit Log용)
    old_category = db.query(SafetyCategory).filter(SafetyCategory.id == category_id).first()
    if old_category is None:
        raise HTTPException(status_code=404, detail="Category not found")
    
    old_data = model_to_dict(old_category)
    
    # 업데이트 수행
    db_category = category_crud.update_category(db, category_id, category)
    
    # Audit Log 기록
    new_data = model_to_dict(db_category)
    log_category_update(db, category_id, old_data, new_data, request)
    
    return db_category

@router.put("/categories/{category_id}/form", response_model=Category)
async def update_category_form(
    category_id: int,
    name: Optional[str] = Form(None),
    code: Optional[str] = Form(None),
    slug: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    display_order: Optional[int] = Form(None),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    """카테고리 정보를 수정합니다 (FormData 지원)."""
    from schemas.category import CategoryUpdate
    
    # CategoryUpdate 객체 생성 (None이 아닌 값들만)
    category_data = {}
    if name is not None and name.strip():
        category_data['name'] = name.strip()
    if code is not None and code.strip():
        category_data['code'] = code.strip()
    if slug is not None and slug.strip():
        category_data['slug'] = slug.strip()
    if description is not None:
        category_data['description'] = description.strip() if description.strip() else None
    if display_order is not None:
        category_data['display_order'] = display_order
    
    if not category_data:
        raise HTTPException(status_code=400, detail="업데이트할 데이터가 없습니다")
    
    category = CategoryUpdate(**category_data)
    db_category = category_crud.update_category(db, category_id, category)
    if db_category is None:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return db_category

@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """카테고리를 삭제합니다."""
    # 기존 데이터 조회 (Audit Log용)
    old_category = db.query(SafetyCategory).filter(SafetyCategory.id == category_id).first()
    if old_category is None:
        raise HTTPException(status_code=404, detail="Category not found")
    
    old_data = model_to_dict(old_category)
    
    # 삭제 수행
    db_category = category_crud.delete_category(db, category_id)
    
    # Audit Log 기록
    log_category_delete(db, category_id, old_data, request)
    
    return {"message": f"카테고리 {category_id}가 성공적으로 삭제되었습니다"}

# ============= 제품 관리 =============

@router.get("/products", response_model=List[ProductResponse])
async def read_products(
    skip: int = 0,
    limit: int = 1000,
    category_code: str = None,
    search: str = None,
    db: Session = Depends(get_db)
):
    """제품 목록을 조회합니다."""
    products = product_crud.get_products(db, skip=skip, limit=limit, category_code=category_code, search=search)
    return products

@router.get("/products/{product_id}", response_model=ProductResponse)
async def read_product(
    product_id: int,
    db: Session = Depends(get_db)
):
    """특정 제품을 조회합니다."""
    product = product_crud.get_product(db, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")
    return product

@router.post("/products", response_model=ProductResponse)
async def create_product(
    product: ProductCreate,
    request: Request,
    db: Session = Depends(get_db)
):
    """새로운 제품을 추가합니다 (JSON 지원)."""
    logger.info(f"Creating product: {product.dict()}")
    created_product = product_crud.create_product(db, product)
    
    # Audit Log 기록
    log_product_create(db, created_product.id, product.dict(), request)
    
    return created_product

@router.post("/products/form", response_model=ProductResponse)
async def create_product_form(
    name: str = Form(...),
    model_number: str = Form(...),
    category_id: int = Form(...),
    description: Optional[str] = Form(None),
    specifications: Optional[str] = Form(None),
    price: Optional[float] = Form(None),
    is_featured: Optional[bool] = Form(False),
    display_order: Optional[int] = Form(0),
    images: List[UploadFile] = File([]),
    db: Session = Depends(get_db)
):
    """새로운 제품을 추가합니다 (FormData 지원)."""
    from schemas.product import ProductCreate
    
    # ProductCreate 객체 생성
    product_data = {
        'name': name.strip() if name else "",
        'model_number': model_number.strip() if model_number else "",
        'category_id': category_id if category_id is not None else 0,
        'description': description.strip() if description and description.strip() else None,
        'specifications': specifications.strip() if specifications and specifications.strip() else None,
        'price': price if price is not None else None,
        'is_featured': is_featured if is_featured is not None else False,
        'display_order': int(display_order) if display_order is not None else 0
    }
    
    # 여러 이미지 파일 처리 및 저장
    image_paths = []
    if images and len(images) > 0:
        import os
        import uuid
        import json
        from pathlib import Path
        
        # 업로드 디렉토리 설정
        upload_dir = Path("../frontend/public/images")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        for image in images:
            if image.filename and image.filename.strip():
                # 고유한 파일명 생성
                file_extension = os.path.splitext(image.filename)[1].lower()
                unique_filename = f"{uuid.uuid4()}{file_extension}"
                
                # 파일 저장
                file_path = upload_dir / unique_filename
                contents = await image.read()
                with open(file_path, "wb") as f:
                    f.write(contents)
                
                image_path = f"/images/{unique_filename}"
                image_paths.append(image_path)
                logger.info(f"이미지 저장 완료: {unique_filename}")
        
        if image_paths:
            # 첫 번째 이미지를 메인 이미지로 설정
            product_data['file_name'] = os.path.basename(image_paths[0])
            # 모든 이미지 경로를 JSON 배열로 저장
            product_data['file_path'] = json.dumps(image_paths)
        else:
            # 기본값 설정
            product_data['file_name'] = "default.jpg"
            product_data['file_path'] = json.dumps(["/images/default.jpg"])
    else:
        # 기본값 설정
        product_data['file_name'] = "default.jpg"
        product_data['file_path'] = json.dumps(["/images/default.jpg"])
    
    try:
        product = ProductCreate(**product_data)
        created_product = product_crud.create_product(db, product)
        return created_product
    except Exception as e:
        logger.error(f"Error creating product: {e}")
        raise HTTPException(status_code=400, detail=f"제품 생성 실패: {str(e)}")
        
        if image_paths:
            # 첫 번째 이미지를 메인 이미지로 설정
            product_data['file_name'] = os.path.basename(image_paths[0])
            # 모든 이미지 경로를 JSON 배열로 저장
            product_data['file_path'] = json.dumps(image_paths)
        else:
            # 기본값 설정
            product_data['file_name'] = "default.jpg"
            product_data['file_path'] = json.dumps(["/images/default.jpg"])
    else:
        # 기본값 설정
        product_data['file_name'] = "default.jpg"
        product_data['file_path'] = json.dumps(["/images/default.jpg"])
    
    print(f"product_data to create: {product_data}")
    
    try:
        product = ProductCreate(**product_data)
        print(f"ProductCreate object: {product}")
        
        created_product = product_crud.create_product(db, product)
        print(f"Created product: {created_product}")
        
        return created_product
    except Exception as e:
        print(f"Error creating ProductCreate object: {e}")
        print(f"product_data that failed: {product_data}")
        raise HTTPException(status_code=400, detail=f"제품 생성 실패: {str(e)}")

@router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    product: ProductUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """기존 제품 정보를 수정합니다 (JSON 지원)."""
    logger.info(f"Updating product {product_id}: {product.dict(exclude_unset=True)}")
    
    # 기존 데이터 조회 (Audit Log용)
    old_product = db.query(SafetyProduct).filter(SafetyProduct.id == product_id).first()
    if old_product is None:
        raise HTTPException(status_code=404, detail="Product not found")
    
    old_data = model_to_dict(old_product)
    
    # 업데이트 수행
    db_product = product_crud.update_product(db, product_id, product)
    
    # Audit Log 기록
    new_data = model_to_dict(db_product)
    log_product_update(db, product_id, old_data, new_data, request)
    
    return db_product

@router.put("/products/{product_id}/form", response_model=ProductResponse)
async def update_product_form(
    product_id: int,
    name: Optional[str] = Form(None),
    model_number: Optional[str] = Form(None),
    category_id: Optional[int] = Form(None),
    description: Optional[str] = Form(None),
    specifications: Optional[str] = Form(None),
    price: Optional[float] = Form(None),
    is_featured: Optional[bool] = Form(None),
    display_order: Optional[int] = Form(None),
    existing_images: Optional[str] = Form(None),  # 유지할 기존 이미지들 (JSON 배열)
    images: List[UploadFile] = File([]),
    db: Session = Depends(get_db)
):
    """기존 제품 정보를 수정합니다 (FormData 지원)."""
    from schemas.product import ProductUpdate
    
    # ProductUpdate 객체 생성 (None이 아닌 값들만)
    product_data = {}
    if name is not None and name.strip():
        product_data['name'] = name.strip()
    if model_number is not None and model_number.strip():
        product_data['model_number'] = model_number.strip()
    if category_id is not None:
        product_data['category_id'] = category_id
    if description is not None:
        product_data['description'] = description.strip() if description.strip() else None
    if specifications is not None:
        product_data['specifications'] = specifications.strip() if specifications.strip() else None
    if price is not None:
        product_data['price'] = price
    if is_featured is not None:
        product_data['is_featured'] = is_featured
    if display_order is not None:
        product_data['display_order'] = display_order
    
    # 이미지 처리 (기존 + 새로운 이미지)
    import os
    import uuid
    import json
    from pathlib import Path
    
    # 기존 제품 정보 가져오기
    existing_product = db.query(SafetyProduct).filter(SafetyProduct.id == product_id).first()
    if not existing_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # 1. 유지할 기존 이미지 파싱
    keep_existing_paths = []
    if existing_images:
        try:
            keep_existing_paths = json.loads(existing_images)
            if not isinstance(keep_existing_paths, list):
                keep_existing_paths = []
        except (json.JSONDecodeError, TypeError):
            keep_existing_paths = []
    
    # 2. 삭제될 기존 이미지 파악 및 실제 파일 삭제
    if existing_product.file_path:
        try:
            all_existing_paths = json.loads(existing_product.file_path)
            if isinstance(all_existing_paths, list):
                paths_to_delete = [path for path in all_existing_paths if path not in keep_existing_paths]
                for path_to_delete in paths_to_delete:
                    if path_to_delete and path_to_delete.startswith('/images/'):
                        file_path = Path("../frontend/public") / path_to_delete.lstrip('/')
                        if file_path.exists():
                            try:
                                os.remove(file_path)
                                logger.info(f"삭제된 이미지 파일: {file_path}")
                            except Exception as e:
                                logger.error(f"이미지 파일 삭제 실패: {file_path}, 오류: {e}")
        except (json.JSONDecodeError, TypeError):
            pass
    
    # 3. 새 이미지들 저장
    new_image_paths = []
    if images and len(images) > 0:
        upload_dir = Path("../frontend/public/images")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        for image in images:
            if image.filename and image.filename.strip():
                # 고유한 파일명 생성
                file_extension = os.path.splitext(image.filename)[1].lower()
                unique_filename = f"{uuid.uuid4()}{file_extension}"
                
                # 파일 저장
                file_path = upload_dir / unique_filename
                contents = await image.read()
                with open(file_path, "wb") as f:
                    f.write(contents)
                
                image_path = f"/images/{unique_filename}"
                new_image_paths.append(image_path)
                logger.info(f"새 이미지 저장 완료: {unique_filename}")
    
    # 4. 최종 이미지 경로 리스트 구성 (기존 유지 + 새 이미지)
    final_image_paths = keep_existing_paths + new_image_paths
    
    if final_image_paths:
        product_data['file_name'] = os.path.basename(final_image_paths[0])
        product_data['file_path'] = json.dumps(final_image_paths)
    elif existing_product.file_path:  # 모든 이미지가 삭제된 경우
        product_data['file_name'] = "default.jpg"
        product_data['file_path'] = json.dumps(["/images/default.jpg"])
    
    if not product_data:
        raise HTTPException(status_code=400, detail="업데이트할 데이터가 없습니다")
    
    try:
        product = ProductUpdate(**product_data)
        db_product = product_crud.update_product(db, product_id, product)
        if db_product is None:
            raise HTTPException(status_code=404, detail="Product not found")
        return db_product
    except ValueError as ve:
        raise HTTPException(status_code=422, detail=f"데이터 검증 오류: {str(ve)}")
    except Exception as e:
        logger.error(f"Error updating product: {e}")
        raise HTTPException(status_code=500, detail=f"제품 업데이트 실패: {str(e)}")

@router.delete("/products/{product_id}")
async def delete_product(
    product_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """제품을 삭제합니다."""
    # 기존 데이터 조회 (Audit Log용)
    old_product = db.query(SafetyProduct).filter(SafetyProduct.id == product_id).first()
    if old_product is None:
        raise HTTPException(status_code=404, detail="Product not found")
    
    old_data = model_to_dict(old_product)
    
    # 삭제 수행
    db_product = product_crud.delete_product(db, product_id)
    
    # Audit Log 기록
    log_product_delete(db, product_id, old_data, request)
    
    return {"message": f"제품 {product_id}가 성공적으로 삭제되었습니다"}

@router.post("/products/{product_id}/duplicate", response_model=ProductResponse)
async def duplicate_product(
    product_id: int,
    db: Session = Depends(get_db)
):
    """기존 제품을 복사하여 새 제품을 생성합니다."""
    import json
    
    # 원본 제품 조회
    original = db.query(SafetyProduct).filter(SafetyProduct.id == product_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # 새 제품 데이터 생성 (ID 제외)
    new_product_data = {
        "name": f"{original.name} (복사본)",
        "model_number": f"{original.model_number}-COPY",
        "category_id": original.category_id,
        "description": original.description,
        "specifications": original.specifications,
        "price": original.price,
        "stock_status": original.stock_status,
        "is_featured": 0,  # 복사본은 기본적으로 추천 해제
        "display_order": original.display_order,
        "file_name": original.file_name,
        "file_path": original.file_path
    }
    
    # 새 제품 생성
    new_product = SafetyProduct(**new_product_data)
    db.add(new_product)
    db.commit()
    db.refresh(new_product)
    
    logger.info(f"제품 복사 완료: {product_id} → {new_product.id}")
    return product_crud.get_product(db, new_product.id)

@router.put("/products/bulk")
async def bulk_update_products(
    product_ids: List[int],
    updates: dict,
    db: Session = Depends(get_db)
):
    """여러 제품을 일괄 수정합니다."""
    from schemas.product import ProductUpdate
    
    if not product_ids:
        raise HTTPException(status_code=400, detail="제품 ID 목록이 비어있습니다")
    
    # 허용된 필드만 업데이트
    allowed_fields = ['category_id', 'price', 'is_featured', 'stock_status', 'display_order']
    filtered_updates = {k: v for k, v in updates.items() if k in allowed_fields}
    
    if not filtered_updates:
        raise HTTPException(status_code=400, detail="업데이트할 필드가 없습니다")
    
    # is_featured 타입 변환
    if 'is_featured' in filtered_updates:
        filtered_updates['is_featured'] = 1 if filtered_updates['is_featured'] else 0
    
    # 일괄 업데이트 실행
    updated_count = 0
    for product_id in product_ids:
        product = db.query(SafetyProduct).filter(SafetyProduct.id == product_id).first()
        if product:
            for key, value in filtered_updates.items():
                setattr(product, key, value)
            updated_count += 1
    
    db.commit()
    
    logger.info(f"일괄 수정 완료: {updated_count}개 제품")
    return {
        "message": f"{updated_count}개 제품이 성공적으로 수정되었습니다",
        "updated_count": updated_count,
        "total_requested": len(product_ids)
    }

@router.delete("/products/bulk")
async def bulk_delete_products(
    product_ids: List[int],
    db: Session = Depends(get_db)
):
    """여러 제품을 일괄 삭제합니다."""
    if not product_ids:
        raise HTTPException(status_code=400, detail="제품 ID 목록이 비어있습니다")
    
    deleted_count = 0
    for product_id in product_ids:
        product = product_crud.delete_product(db, product_id)
        if product:
            deleted_count += 1
    
    logger.info(f"일괄 삭제 완료: {deleted_count}개 제품")
    return {
        "message": f"{deleted_count}개 제품이 성공적으로 삭제되었습니다",
        "deleted_count": deleted_count,
        "total_requested": len(product_ids)
    }

@router.get("/products/export/template")
async def download_product_template():
    """제품 등록용 엑셀 템플릿을 다운로드합니다."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from io import BytesIO
    
    # 엑셀 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "제품 목록"
    
    # 헤더 작성
    headers = [
        "제품명*", "모델번호*", "카테고리코드*", "설명", "규격",
        "가격", "재고상태", "추천제품", "표시순서"
    ]
    ws.append(headers)
    
    # 예시 데이터
    example_row = [
        "안전모 예시", "SH-001", "safety_helmet", "고급 안전모",
        "ABS 재질, 조절식", 25000, "in_stock", "FALSE", 1
    ]
    ws.append(example_row)
    
    # 스타일 적용
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # BytesIO로 저장
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_template.xlsx"}
    )

@router.post("/products/import")
async def import_products_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """엑셀 파일로 제품을 일괄 등록합니다."""
    import openpyxl
    from io import BytesIO
    import json
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="엑셀 파일만 업로드 가능합니다 (.xlsx, .xls)")
    
    # 파일 읽기
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    
    # 헤더 검증 (첫 번째 행)
    headers = [cell.value for cell in ws[1]]
    required_headers = ["제품명", "모델번호", "카테고리코드"]
    
    # 결과 저장
    results = {
        "success": [],
        "errors": [],
        "total": 0
    }
    
    # 데이터 행 처리 (2행부터)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        results["total"] += 1
        
        try:
            # 빈 행 건너뛰기
            if not any(row):
                continue
            
            # 데이터 파싱
            name = row[0]
            model_number = row[1]
            category_code = row[2]
            description = row[3] if len(row) > 3 else None
            specifications = row[4] if len(row) > 4 else None
            price = float(row[5]) if len(row) > 5 and row[5] else None
            stock_status = row[6] if len(row) > 6 else "in_stock"
            is_featured = str(row[7]).upper() == "TRUE" if len(row) > 7 else False
            display_order = int(row[8]) if len(row) > 8 and row[8] else 0
            
            # 필수 필드 검증
            if not name or not model_number or not category_code:
                results["errors"].append({
                    "row": row_idx,
                    "error": "필수 필드 누락 (제품명, 모델번호, 카테고리코드)"
                })
                continue
            
            # 카테고리 존재 확인
            category = category_crud.get_category_by_code(db, category_code)
            if not category:
                results["errors"].append({
                    "row": row_idx,
                    "error": f"존재하지 않는 카테고리 코드: {category_code}"
                })
                continue
            
            # 제품 생성
            product_data = {
                "name": name,
                "model_number": model_number,
                "category_id": category.id,
                "description": description,
                "specifications": specifications,
                "price": price,
                "stock_status": stock_status,
                "is_featured": 1 if is_featured else 0,
                "display_order": display_order,
                "file_name": "default.jpg",
                "file_path": json.dumps(["/images/default.jpg"])
            }
            
            new_product = SafetyProduct(**product_data)
            db.add(new_product)
            db.commit()
            db.refresh(new_product)
            
            results["success"].append({
                "row": row_idx,
                "id": new_product.id,
                "name": name
            })
            
        except Exception as e:
            results["errors"].append({
                "row": row_idx,
                "error": str(e)
            })
            db.rollback()
    
    logger.info(f"엑셀 업로드 완료: 성공 {len(results['success'])}개, 실패 {len(results['errors'])}개")
    
    return {
        "message": f"총 {results['total']}개 중 {len(results['success'])}개 성공, {len(results['errors'])}개 실패",
        "success_count": len(results["success"]),
        "error_count": len(results["errors"]),
        "details": results
    }

@router.post("/images/bulk")
async def bulk_upload_images(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db)
):
    """여러 이미지를 동시에 업로드합니다."""
    import os
    import uuid
    from pathlib import Path
    
    if not files:
        raise HTTPException(status_code=400, detail="업로드할 파일이 없습니다")
    
    upload_dir = Path("../frontend/public/images")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    results = {
        "success": [],
        "errors": []
    }
    
    for file in files:
        try:
            # 파일 형식 검증
            allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
            file_extension = os.path.splitext(file.filename)[1].lower()
            
            if file_extension not in allowed_extensions:
                results["errors"].append({
                    "filename": file.filename,
                    "error": "지원하지 않는 파일 형식"
                })
                continue
            
            # 파일 크기 검증 (10MB 제한)
            contents = await file.read()
            if len(contents) > 10 * 1024 * 1024:  # 10MB
                results["errors"].append({
                    "filename": file.filename,
                    "error": "파일 크기가 너무 큽니다 (최대 10MB)"
                })
                continue
            
            # 고유한 파일명 생성
            unique_filename = f"{uuid.uuid4()}{file_extension}"
            file_path = upload_dir / unique_filename
            
            # 파일 저장
            with open(file_path, "wb") as f:
                f.write(contents)
            
            results["success"].append({
                "original_filename": file.filename,
                "saved_filename": unique_filename,
                "url": f"/images/{unique_filename}",
                "size": len(contents)
            })
            
            logger.info(f"이미지 업로드 성공: {file.filename} → {unique_filename}")
            
        except Exception as e:
            results["errors"].append({
                "filename": file.filename,
                "error": str(e)
            })
            logger.error(f"이미지 업로드 실패: {file.filename}, 오류: {e}")
    
    return {
        "message": f"총 {len(files)}개 중 {len(results['success'])}개 성공, {len(results['errors'])}개 실패",
        "success_count": len(results["success"]),
        "error_count": len(results["errors"]),
        "results": results
    }

@router.post("/content")
async def create_content(content: dict):
    """새로운 컨텐츠를 추가하는 엔드포인트"""
    return {
        "message": "컨텐츠가 성공적으로 추가되었습니다",
        "content": content
    }

@router.put("/content/{content_id}")
async def update_content(content_id: str, content: dict):
    """기존 컨텐츠를 수정하는 엔드포인트"""
    return {
        "message": f"컨텐츠 {content_id}가 성공적으로 수정되었습니다",
        "content": content
    }

@router.delete("/content/{content_id}")
async def delete_content(content_id: str):
    """컨텐츠를 삭제하는 엔드포인트"""
    return {
        "message": f"컨텐츠 {content_id}가 성공적으로 삭제되었습니다"
    }


# ============================================================
# 고급 검색 & 필터링
# ============================================================

@router.post("/products/advanced-search", response_model=ProductSearchResponse)
async def admin_advanced_search_products(
    params: ProductSearchParams,
    db: Session = Depends(get_db)
):
    """
    관리자용 고급 검색으로 제품을 조회합니다.
    
    - **search**: 텍스트 검색 (제품명, 설명, 모델번호, 사양)
    - **category_id**: 단일 카테고리 ID
    - **category_codes**: 여러 카테고리 코드 (예: ["safety_helmet", "safety_gloves"])
    - **min_price / max_price**: 가격 범위
    - **stock_status**: 재고 상태 ("재고있음", "품절", "입고예정" 등)
    - **is_featured**: 추천 제품 여부
    - **created_after / created_before**: 등록 날짜 범위
    - **sort_by**: 정렬 필드 (name, price, created_at, updated_at, display_order)
    - **sort_order**: 정렬 순서 (asc, desc)
    - **skip / limit**: 페이징
    
    예시:
    ```json
    {
        "search": "헬멧",
        "category_codes": ["safety_helmet"],
        "min_price": 10000,
        "max_price": 50000,
        "stock_status": "재고있음",
        "is_featured": true,
        "sort_by": "price",
        "sort_order": "asc",
        "skip": 0,
        "limit": 20
    }
    ```
    """
    products, total = product_crud.advanced_search_products(db, params)
    
    # 페이지 정보 계산
    page = (params.skip // params.limit) + 1 if params.limit > 0 else 1
    total_pages = math.ceil(total / params.limit) if params.limit > 0 else 0
    
    return ProductSearchResponse(
        total=total,
        items=products,
        page=page,
        page_size=params.limit,
        total_pages=total_pages
    )


# ============================================================
# 변경 이력 추적 (Audit Log)
# ============================================================

@router.get("/audit-logs", response_model=List[AuditLogResponse])
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """
    변경 이력 로그를 조회합니다.
    
    - **entity_type**: 엔티티 타입 (PRODUCT, CATEGORY)
    - **entity_id**: 엔티티 ID
    - **action**: 작업 타입 (CREATE, UPDATE, DELETE, BULK_UPDATE, BULK_DELETE)
    - **user_id**: 사용자 ID (향후 인증 시스템 추가 시 사용)
    - **date_from**: 시작 날짜 (YYYY-MM-DD 또는 ISO 8601 형식)
    - **date_to**: 종료 날짜
    - **skip / limit**: 페이징
    """
    from datetime import datetime
    
    # 날짜 문자열을 datetime으로 변환
    date_from_dt = None
    date_to_dt = None
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from)
        except:
            pass
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to)
        except:
            pass
    
    # 필터 생성
    filters = AuditLogFilter(
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        user_id=user_id,
        date_from=date_from_dt,
        date_to=date_to_dt,
        skip=skip,
        limit=limit
    )
    
    logs, total = audit_crud.get_audit_logs(db, filters)
    return logs


@router.get("/audit-logs/entity/{entity_type}/{entity_id}", response_model=List[AuditLogResponse])
async def get_entity_audit_history(
    entity_type: str,
    entity_id: int,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """
    특정 엔티티(제품/카테고리)의 변경 이력을 조회합니다.
    
    예: GET /api/admin/audit-logs/entity/PRODUCT/123
    """
    entity_type_enum = AuditEntityType(entity_type)
    logs = audit_crud.get_entity_history(db, entity_type_enum, entity_id, limit)
    return logs


@router.get("/audit-logs/recent", response_model=List[AuditLogResponse])
async def get_recent_audit_activities(
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """최근 변경 활동을 조회합니다 (대시보드용)."""
    logs = audit_crud.get_recent_activities(db, limit)
    return logs

 